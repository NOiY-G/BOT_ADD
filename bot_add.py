import time
from telethon.sync import TelegramClient
from telethon.tl.functions.channels import InviteToChannelRequest
from telethon.tl.types import InputPeerChannel
from openpyxl import load_workbook

# Credenciais da API de administrador
api_id = 'COLOQUE SUA API ID'
api_hash = 'COLOQUE SUA API HASH'

# ID do seu canal do Telegram
channel_username = 'COLOQUE O ID DO CANAL (@ DO CANAL)'

# Nome do arquivo Excel contendo os IDs dos usuários
excel_file = 'COLOQUE O CAMINHO DO ARQUIVO XLSX'   

# Função para ler IDs do Excel e adicionar ao canal
def adicionar_membros_do_excel(delay=10):
    client = TelegramClient('session_name', api_id, api_hash)

    with client:
        # Abrir o arquivo Excel
        wb = load_workbook(excel_file)
        sheet = wb.active

        for row in sheet.iter_rows(min_row=2, values_only=True):  # Começa da segunda linha para ignorar cabeçalhos
            user_id = row[0]  # Supondo que o ID do usuário está na primeira coluna
            if user_id:
                try:
                    # Obter o InputPeer do canal
                    channel = client.get_entity(channel_username)
                    channel_peer = InputPeerChannel(channel_id=channel.id, access_hash=channel.access_hash)

                    # Adicionar usuário ao canal
                    client(InviteToChannelRequest(channel_peer, [user_id]))

                    print(f'Usuário {user_id} adicionado ao canal {channel_username} com sucesso!')

                    # Aguardar um intervalo de tempo (delay) antes de adicionar o próximo usuário
                    time.sleep(delay)

                except Exception as e:
                    print(f'Erro ao adicionar usuário {user_id} ao canal: {str(e)}')

        wb.close()

# Chamada da função principal com um delay de 10 segundos entre cada adição
adicionar_membros_do_excel(delay=10)
